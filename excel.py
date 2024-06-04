import os
import argparse
from openpyxl import Workbook, load_workbook, styles

parser = argparse.ArgumentParser()
parser.add_argument('--path')
args = parser.parse_args()

target_path = args.path

def read_data(text_path):
    data = []
    with open(text_path, 'r') as file:
        for line in file:
            line = line.strip()
            if line=='psnr' or line=='msssim' or line=='lpips' or line=='orig' or line=='quant':
                continue
            elif line == '':
                continue
            else:
                parts = line.split(':')
                parts = [part.strip() for part in parts]
                data.append(parts[1])
    return data

eval_list = []
for folder in os.listdir(target_path):
    folder_path = os.path.join(target_path, folder)
    for f in os.listdir(folder_path):
        f_path = os.path.join(folder_path, f)
        if os.path.isdir(f_path):
            if f_path.endswith('eval') and not f_path.endswith('hnerv_eval'):
                eval_list.append(f_path)
eval_list.sort()
print(eval_list)

ws_list = ['PSNR', 'MSSSIM', 'LPIPS']
ws_list2 = ['orig', 'quant']
ws_list3 = ['Proposed', 'HNeRV']
wb = Workbook()
save_path = os.path.join(args.path, 'compare.xlsx')
wb.save(save_path)
wb = load_workbook(save_path)
ws = wb['Sheet']
ws = wb.active
ws.cell(row=1, column=1, value=os.path.basename(args.path))
ws.cell(row=3, column=2, value='Judge')
ws.cell(row=2, column=8, value='Evaluation')
ws.cell(row=3, column=8, value='orig')
ws.cell(row=3, column=14, value='quant')
for l in range(3):
    for i in range(len(ws_list)):
        ws.cell(row=5, column=2+i+(l*3), value=ws_list[i])
for l in range(2):
    for i in range(len(ws_list)):
        ws.cell(row=4, column=8+(i*2)+(l*6), value=ws_list[i])
for i in range(len(ws_list2)):
    ws.cell(row=4, column=2+(i*3), value=ws_list2[i])
for l in range(6):
    for i in range(len(ws_list3)):
        ws.cell(row=5, column=8+i+(l*2), value=ws_list3[i])
wb.save(save_path)

idx_list = [12, 13, 14, 15, 16, 17, 0, 1, 4, 5, 8, 9, 2, 3, 6, 7, 10, 11]
for i in range(len(eval_list)):
    folder = eval_list[i]
    text_path = os.path.join(folder, 'compare.txt')
    data = read_data(text_path)
    sorted_data = [data[i] for i in idx_list]
    name = os.path.basename(folder)
    names = name.split('_')
    ws.cell(row=6+i, column=1, value=names[0])
    for l in range(len(sorted_data)):
        if sorted_data[l] == 'Propo':
            ws.cell(row=6+i, column=2+l, value=1)
        elif sorted_data[l] == 'HNeRV':
            ws.cell(row=6+i, column=2+l, value=0).font=styles.fonts.Font(color='FF0000')
        else:
            ws.cell(row=6+i, column=2+l, value=float(sorted_data[l]))
    wb.save(save_path)

length = len(eval_list)
for i in range(6):
    sum = 0
    for j in range(length):
        sum += int(ws.cell(row=6+j, column=2+i).value)
    ws.cell(row=6+length, column=2+i, value=f'{str(sum)}/{str(length)}')
    wb.save(save_path)
wb.save(save_path)